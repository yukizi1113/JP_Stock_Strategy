"""
データローダー
stooq / yfinance / JPX xls からデータを取得する共通ユーティリティ。

OHLCV（始値・高値・安値・終値・出来高）データも取得できるよう拡張。
JQuants-Forum の知見に基づきATR・HLバンド・出来高比率特徴量の計算も含む。
"""
from __future__ import annotations
import io
import re
import warnings
import requests
import numpy as np
import pandas as pd
import yfinance as yf
import pandas_datareader.data as web
from datetime import date, timedelta
from typing import List, Optional, Dict

from config import JPX_DATA_URL, JPX_EXCLUDE_TYPES, PRICE_SUFFIX

warnings.filterwarnings("ignore", category=FutureWarning)


# ─────────────────────────── JPX ユニバース ───────────────────────────────

_NO_SHORT_CACHE: Optional[frozenset] = None

JPX_MARGIN_URL = (
    "https://www.jpx.co.jp/listing/others/margin/"
    "tvdivq0000000od2-att/20260202_list.xlsx"
)


def build_no_short_tickers(
    url: str = JPX_MARGIN_URL,
    refresh: bool = False,
) -> frozenset:
    """
    JPX貸借銘柄リストを取得し、空売り不可銘柄のtickerセットを返す。

    日本市場では「制度信用銘柄」「非制度信用銘柄」は空売りできない。
    空売り可能なのは「貸借銘柄」のみ。

    Parameters
    ----------
    url     : JPX貸借銘柄リストExcelファイルのURL（定期的に更新要）
    refresh : True の場合キャッシュを無視して再取得

    Returns
    -------
    frozenset of str : 空売り不可ティッカー（4桁コード）の集合
    """
    global _NO_SHORT_CACHE
    if _NO_SHORT_CACHE is not None and not refresh:
        return _NO_SHORT_CACHE

    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl が必要です: pip install openpyxl")

    headers = {"User-Agent": "Mozilla/5.0"}
    r = requests.get(url, timeout=30, headers=headers)
    r.raise_for_status()

    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    sheet_name = "一覧" if "一覧" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    NO_SHORT_VALUES = {"制度信用銘柄", "非制度信用銘柄"}
    no_short: set[str] = set()

    # 行1: 空白ヘッダー, 行2: 実際のヘッダー, 行3以降: データ
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx <= 2:
            continue
        if not row or row[0] is None:
            continue

        code   = str(row[0]).strip()
        shinyo = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""

        if shinyo in NO_SHORT_VALUES:
            if code.endswith(".0"):
                code = code[:-2]
            no_short.add(code)

    _NO_SHORT_CACHE = frozenset(no_short)
    return _NO_SHORT_CACHE


def build_jpx_universe(exclude_types: frozenset = JPX_EXCLUDE_TYPES) -> pd.DataFrame:
    """
    JPXの data_j.xls を取得し、上場普通株式のDataFrameを返す。
    Columns: ticker, company, market_type
    """
    try:
        import xlrd
    except ImportError:
        raise ImportError("xlrd が必要です: pip install xlrd")

    r = requests.get(JPX_DATA_URL, timeout=30)
    r.raise_for_status()
    wb = xlrd.open_workbook(file_contents=r.content)
    ws = wb.sheet_by_index(0)

    rows = []
    for i in range(1, ws.nrows):
        raw_code = ws.cell_value(i, 1)   # B列: コード
        company  = ws.cell_value(i, 2)   # C列: 銘柄名
        mtype    = str(ws.cell_value(i, 3)).strip()  # D列: 市場区分

        if mtype in exclude_types:
            continue

        raw_str = str(raw_code).strip()
        if raw_str.endswith(".0"):
            raw_str = raw_str[:-2]
        ticker = re.sub(r"0$", "", raw_str) if re.match(r"^\d+$", raw_str) else raw_str

        rows.append({"ticker": ticker, "company": str(company).strip(), "market_type": mtype})

    return pd.DataFrame(rows)


# ─────────────────────────── OHLCV データ取得 ────────────────────────────

def fetch_ohlcv_yfinance(
    tickers: List[str],
    start: str,
    end: str,
) -> Dict[str, pd.DataFrame]:
    """
    yfinance で複数銘柄のOHLCVデータを取得する。
    Returns: {ticker: DataFrame(index=date, columns=[Open,High,Low,Close,Volume])}
    """
    yf_tickers = [f"{tk}{PRICE_SUFFIX}" for tk in tickers]
    raw = yf.download(yf_tickers, start=start, end=end, auto_adjust=True, progress=False)
    if raw.empty:
        return {}

    raw.index = pd.to_datetime(raw.index).tz_localize(None)
    result = {}

    if len(tickers) == 1:
        df = raw[["Open", "High", "Low", "Close", "Volume"]].copy()
        df.columns = ["Open", "High", "Low", "Close", "Volume"]
        result[tickers[0]] = df
    else:
        for tk, yfk in zip(tickers, yf_tickers):
            try:
                df = pd.DataFrame({
                    "Open":   raw["Open"][yfk],
                    "High":   raw["High"][yfk],
                    "Low":    raw["Low"][yfk],
                    "Close":  raw["Close"][yfk],
                    "Volume": raw["Volume"][yfk],
                })
                df.index = pd.to_datetime(df.index).tz_localize(None)
                result[tk] = df.dropna(how="all")
            except Exception:
                pass

    return result


def fetch_prices_stooq(
    tickers: List[str],
    start: str,
    end: str,
    field: str = "Close",
) -> pd.DataFrame:
    """pandas-datareader (stooq) で複数銘柄の終値を取得"""
    frames = {}
    for tk in tickers:
        symbol = f"{tk}.JP"
        try:
            df = web.DataReader(symbol, "stooq", start=start, end=end)
            if field in df.columns and not df.empty:
                frames[tk] = df[field].sort_index()
        except Exception:
            pass
    if not frames:
        return pd.DataFrame()
    return pd.DataFrame(frames).sort_index()


def fetch_prices_yfinance(
    tickers: List[str],
    start: str,
    end: str,
    field: str = "Close",
) -> pd.DataFrame:
    """yfinance で複数銘柄の終値を取得"""
    yf_tickers = [f"{tk}{PRICE_SUFFIX}" for tk in tickers]
    raw = yf.download(yf_tickers, start=start, end=end, auto_adjust=True, progress=False)
    if raw.empty:
        return pd.DataFrame()
    if len(tickers) == 1:
        prices = raw[[field]].rename(columns={field: tickers[0]})
    else:
        prices = raw[field].copy()
        prices.columns = [c.replace(PRICE_SUFFIX, "") for c in prices.columns]
    prices.index = pd.to_datetime(prices.index).tz_localize(None)
    return prices.sort_index()


def fetch_prices(
    tickers: List[str],
    start: str,
    end: Optional[str] = None,
    source: str = "yfinance",
) -> pd.DataFrame:
    """統一インターフェース: source="stooq" or "yfinance" """
    end = end or date.today().isoformat()
    if source == "stooq":
        return fetch_prices_stooq(tickers, start, end)
    return fetch_prices_yfinance(tickers, start, end)


# ─────────────────────────── JQuants系 特徴量計算 ──────────────────────────

def compute_atr_features(
    ohlcv: pd.DataFrame,
    periods: List[int] = [5, 10, 20, 40, 60],
) -> pd.DataFrame:
    """
    JQuants-Forum 由来のATR 4変種と HL Band を計算する。

    特徴量:
    - atr_{n}:      通常ATR（前日終値・高値・安値考慮）のN日移動平均
    - g_atr_{n}:    ギャップATR（寄付vs前日終値）のN日移動平均
    - d_atr_{n}:    日中ATR（高値-安値）のN日移動平均
    - h_atr_{n}:    ヒゲATR（上下ヒゲ幅）のN日移動平均
    - hl_{n}:       HL Band（N日高値 - N日安値）/ 終値
    - vol_ratio_{n}: N日出来高比率（当日÷N日平均）
    - mi_{n}:       マーケットインパクト（ATR / 出来高比率）
    """
    if ohlcv.empty or "Close" not in ohlcv.columns:
        return pd.DataFrame()

    close  = ohlcv["Close"]
    prev_c = close.shift(1)

    # ATR 4変種（正規化: 前日終値で除す）
    atr_raw = pd.DataFrame({
        "atr":   (ohlcv["High"].combine(prev_c, max) - ohlcv["Low"].combine(prev_c, min)) / prev_c,
        "g_atr": (ohlcv["Open"] - prev_c).abs() / prev_c,
        "d_atr": (ohlcv["High"] - ohlcv["Low"]) / prev_c,
        "h_atr": ((ohlcv["High"] - ohlcv["Low"]) - (ohlcv["Open"] - close).abs()) / prev_c,
    }).clip(lower=0)

    result = {}
    volume = ohlcv["Volume"].replace(0, np.nan)

    for n in periods:
        for col in ["atr", "g_atr", "d_atr", "h_atr"]:
            result[f"{col}_{n}"] = atr_raw[col].rolling(n).mean()

        # HL Band (期間高値-安値レンジ)
        result[f"hl_{n}"] = (
            ohlcv["High"].rolling(n).max() - ohlcv["Low"].rolling(n).min()
        ) / close

        # 出来高比率
        vol_ma = volume.rolling(n).mean()
        result[f"vol_ratio_{n}"] = volume / vol_ma

        # マーケットインパクト
        result[f"mi_{n}"] = atr_raw["atr"].rolling(n).mean() / (vol_ma * close / 1e6).replace(0, np.nan)

    return pd.DataFrame(result, index=ohlcv.index)


def period_wise_normalize(
    df: pd.DataFrame,
    period: str = "QE",
) -> pd.DataFrame:
    """
    期間別正規化 (JQuants-Forum 知見):
    各四半期（または指定期間）内で (x - mean) / std を計算。
    ラベル分布の非定常性を緩和し、モデルの汎化性を向上させる。

    Parameters
    ----------
    df : 正規化対象DataFrame
    period : 集計期間（'QE'=四半期末, 'ME'=月末）
    """
    result = df.copy()
    groups = df.groupby(df.index.to_period(period))
    for _, grp in groups:
        mu  = grp.mean()
        std = grp.std().replace(0, np.nan)
        result.loc[grp.index] = (grp - mu) / std
    return result


# ─────────────────────────── リターン計算 ─────────────────────────────────

def daily_returns(prices: pd.DataFrame) -> pd.DataFrame:
    return prices.pct_change().dropna(how="all")


def log_returns(prices: pd.DataFrame) -> pd.DataFrame:
    return np.log(prices / prices.shift(1)).dropna(how="all")


def monthly_returns(prices: pd.DataFrame) -> pd.DataFrame:
    monthly = prices.resample("ME").last()
    return monthly.pct_change().dropna(how="all")


# ─────────────────────────── ベンチマーク ──────────────────────────────────

def fetch_benchmark(
    start: str,
    end: Optional[str] = None,
    ticker: str = "^N225",
) -> pd.Series:
    end = end or date.today().isoformat()
    raw = yf.download(ticker, start=start, end=end, auto_adjust=True, progress=False)
    s = raw["Close"].squeeze()
    s.index = pd.to_datetime(s.index).tz_localize(None)
    s.name = "benchmark"
    return s


# ─────────────────────────── バックテスト指標 ──────────────────────────────

def performance_stats(
    returns: pd.Series,
    benchmark: Optional[pd.Series] = None,
    rf: float = 0.001,
) -> dict:
    """共通パフォーマンス指標を計算する"""
    r = returns.dropna()
    ann_ret  = (1 + r).prod() ** (252 / len(r)) - 1
    ann_vol  = r.std() * np.sqrt(252)
    sharpe   = (ann_ret - rf) / ann_vol if ann_vol > 0 else np.nan

    cum      = (1 + r).cumprod()
    drawdown = cum / cum.cummax() - 1
    max_dd   = drawdown.min()

    stats = {
        "ann_return":   round(ann_ret, 4),
        "ann_vol":      round(ann_vol, 4),
        "sharpe":       round(sharpe, 4),
        "max_drawdown": round(max_dd, 4),
        "calmar":       round(ann_ret / abs(max_dd), 4) if max_dd != 0 else np.nan,
    }

    if benchmark is not None:
        b = benchmark.reindex(r.index).pct_change().dropna()
        r_aligned = r.reindex(b.index).dropna()
        b_aligned = b.reindex(r_aligned.index)
        if len(b_aligned) > 2:
            beta = np.cov(r_aligned, b_aligned)[0, 1] / np.var(b_aligned)
            alpha = ann_ret - rf - beta * ((1 + b_aligned.mean()) ** 252 - 1 - rf)
            stats["beta"]  = round(float(beta), 4)
            stats["alpha"] = round(float(alpha), 4)

    return stats
